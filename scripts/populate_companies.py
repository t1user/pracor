"""Create database entries for reviews.Company from excel files. Excel files must have labels in row 8 and 1000 entries starting from row 9. Columns with names as per lines 114-123 must exist. The way to use the script is to invoke run_files(function), where function will be performed on every row in every file.
"""

import getpass

# Configuration to allow for access to project models
import os
import sys

import openpyxl
import requests

# from openpyxl.worksheet.worksheet.Worksheet import iter_rows, iter_cells, cell


user = getpass.getuser()

proj_path = "/home/" + user + "/pracor/"

sys.path.append(proj_path)
from django.db import IntegrityError

import pracor.wsgi
from reviews.models import Company

INTEGRITY_COUNT = 0


def get_files():
    """Create a list of excel files (.xlsx extension only) in the current directory.
    All of them will be processed."""
    files = []
    for file in os.listdir():
        if file.endswith(".xlsx"):
            files.append(file)
            files.sort()
    return files


def get_sheet(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    return sheet


def count_rows():
    for file in files:
        sheet = get_sheet(file)
        n = 0
        for row in sheet.rows:
            n += 1
        print("Number of rows in file: ", n)


def get_header(sheet):
    # get content of row 8 and return it as a list - it must contain field
    # names
    header = [
        cell.value
        for column in sheet.iter_cols(min_row=8, max_row=8)
        for cell in column
    ]
    return header


def get_content(header, sheet):
    content = []
    for row in sheet.iter_rows(min_col=1, max_col=len(header), min_row=9, max_row=1008):
        values = {}
        for index, cell in enumerate(row):
            values[header[index]] = cell.value
        # rows without name are empty and should be weeded
        if values["Firma"] is not None:
            content.append(values)
    return content


def create_entry(entry):
    """Creates one database entry from a dictionary passed to it."""

    def get_public(item):
        """Helper function to convert public/private staus into boolean
        value defined in the database"""
        if item == "Niegiełdowe":
            return False
        else:
            return True

    def get_employment(employees):
        """Helper function to convert employment classes into ranges
        defined in database."""
        if employees == "n/a":
            return None
        elif employees < 100:
            return "A"
        elif employees < 500:
            return "B"
        elif employees < 1000:
            return "C"
        elif employees < 5000:
            return "D"
        elif employees < 10000:
            return "E"
        else:
            return "F"

    def clean_name(name):
        new_name = name.strip("[1]").strip()
        if "Likwidacja" in new_name or "Zamknięta" in new_name:
            raise ValueError
        return new_name

    def clean_ownership(owner):
        """
        If the content is too long for database max_length limit - truncate.
        CURRENTLY NOT IN USE. Not required because TextField doesn't have max_length limitation
        (as opposed to CharField).
        """
        return owner
        if len(owner) > 500:
            owner = owner[:499]
        return owner

    def clean_sectors(sectors):
        """
        Same as previous function.
        """
        return sectors
        if len(sectors) > 350:
            sectors = sectors[:349]
        return sectors

    # want to have companies without www removed silently (without loggin an
    # error)
    if www_available(entry):  # and test_www(entry):
        try:
            item = Company(
                name=clean_name(entry["Firma"]),
                headquarters_city=entry["Miasto"],
                region=entry["Kraj/Region"],
                country=entry["Kraj"],
                website=entry["Strona www"],
                public=get_public(entry["Notowane/Nienotowane"]),
                ownership=clean_ownership(entry["Właściciele"]),
                employment=get_employment(entry["Liczba zatrudnionych"]),
                sectors=clean_sectors(entry["Sektory"]),
                isin=entry["ISIN"],
            )
            item.save()
            return True
        except ValueError as e:
            # print('!FIRMA W LIKWIDACJI!', entry['Firma'], entry['Strona www'], 'Error: ', e.args)
            return False
        except IntegrityError as e:
            # print('Integrity error', entry['Firma'], e)
            global INTEGRITY_COUNT
            INTEGRITY_COUNT += 1
        except Exception as e:
            # print('ERROR!!!!!!!!   ', entry['Firma'], entry['Strona www'], 'Error: ', e.args)
            # return False
            print(entry["Firma"], e.args, type(e).__name__)
            sys.exit()
    return False


def test_www(entry):
    """
    Test if the given www is valid. This function can be called inside create_entry (uncomment the call) or a script from separate module can be used.
    """
    number = entry["Num"]
    name = entry["Firma"]
    www = entry["Strona www"]
    try:
        r = requests.get(www)
        r.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print("Error http: ", e.args)
        return False
    except requests.exceptions.ConnectionError as e:
        print("Wrong www: ", e.args)
        return False
    except requests.exceptions.TooManyRedirects as e:
        print("Too many redirects: ", e.args, "still adding ", www, "to database")
        return True
    return True


def www_available(entry):
    if entry["Strona www"] == "n/a":
        # print('No website for: ', entry['Firma'], entry['Strona www'])
        return False
    return True


# returns a list of excel files
files = get_files()
# gets the active sheet from given file
sheet = get_sheet(files[1])
# creates dictionary keys from column labels in row 8, returns a list
header = get_header(sheet)
# returns a list of dictionaries with content from the given sheet
content = get_content(header, sheet)

"""
counter = 0
for entry in content:
    print(entry['Firma'], entry['Strona www'])
    try:
        test_www(entry)
    except:
        counter+=1
print(counter)
"""


def do_file(content, function):
    """Performs 'function' for content generated from one file. Content is a list."""
    counter = 0
    for entry in content:
        a = function(entry)
        counter += a
    return counter


def run_files(function):
    """Performs 'function' for all files."""
    counter = 0
    for file in files:
        print()
        print(
            "------------------------------------------------------------------------------"
        )
        print(file)
        sheet = get_sheet(file)
        header = get_header(sheet)
        content = get_content(header, sheet)
        a = do_file(content, function)
        counter += a
    print("Count: ", counter)


if __name__ == "__main__":
    run_files(create_entry)
    # run_files(list_no_www)
    print("database integrity errors: ", INTEGRITY_COUNT)
